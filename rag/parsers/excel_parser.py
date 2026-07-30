# -*- coding: utf-8 -*-
"""Excel 解析器（.xlsx / .xls）。

逐工作表读入，处理两类结构化难点后再交模型理解：

* 合并单元格：将合并区域的值回填到覆盖的每个单元格，避免语义边界丢失；
* 大表：按行分窗口，每个窗口强制携带表头行，保证数据行不脱离列含义。

每个窗口生成一个表格 Segment，再由基类聚合成 Block。模型不可用时保留 HTML。
"""
from __future__ import annotations

import logging
from io import BytesIO

from . import tables, vlm
from .base import BaseParser, count_tokens
from .doc_types import Block, Segment, SegmentKind

logger = logging.getLogger(__name__)


class ExcelParser(BaseParser):
    """Excel 解析器。"""

    def load(self) -> list[Block]:
        """解析 Excel 文件，逐工作表按窗口生成表格 Segment。"""
        from openpyxl import load_workbook

        workbook = load_workbook(
            BytesIO(self._read_binary()), read_only=False, data_only=True,
        )

        segments: list[Segment] = []
        for worksheet in workbook.worksheets:
            grid = self._sheet_to_matrix(worksheet)
            if not grid:
                continue

            for window in tables.window_rows(grid, header_rows=1):
                table_html = tables.matrix_to_html(window)
                structured = vlm.structure_table(table_html)
                text = f"工作表：{worksheet.title}\n{table_html}"
                segments.append(
                    Segment(
                        text=text,
                        kind=SegmentKind.TABLE,
                        tokens=count_tokens(text),
                        table_html=table_html,
                        payload=structured,
                        section=[worksheet.title],
                    ),
                )

        return self.pack_blocks(segments)

    @staticmethod
    def _sheet_to_matrix(worksheet) -> list[list[str]]:
        """把工作表转为二维字符串矩阵，并展开合并单元格。"""
        raw = [
            ["" if cell is None else str(cell) for cell in row]
            for row in worksheet.iter_rows(values_only=True)
        ]
        # openpyxl 的合并区间是 1 基闭区间，转成 0 基交给工具展开
        ranges = [
            (rng.min_row - 1, rng.min_col - 1, rng.max_row - 1, rng.max_col - 1)
            for rng in worksheet.merged_cells.ranges
        ]
        expanded = tables.expand_merged(raw, ranges)

        # 去掉完全空白的行
        return [row for row in expanded if any(cell.strip() for cell in row)]
