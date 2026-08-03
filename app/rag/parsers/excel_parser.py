# -*- coding: utf-8 -*-
"""Excel（xlsx）解析器。

基于 openpyxl 开源库。
借鉴思路：合并单元格完整展开（来自九阳 POC 复杂表格调优）。
"""
from __future__ import annotations

import logging
from typing import List

from app.rag.parsers.doc_types import ChildChunk, ElementType, ParentChunk
from app.rag.parsers.utils import download_file, is_safe_url
from app.rag.chunkers.smart_chunker import build_child_chunks, build_parent_chunks

logger = logging.getLogger(__name__)

MAX_FILE_SIZE = 50 * 1024 * 1024


class ExcelParser:
    """Excel 文档解析器。

    关键：合并单元格完整展开（九阳 POC 实战，D525 含 533 个合并单元格）。
    """

    def __init__(self, file_uri: str, filename: str):
        if is_safe_url(file_uri):
            self.file_url = file_uri
            self.file_path = None
        else:
            self.file_path = file_uri
            self.file_url = None
        self.filename = filename

    def load(
        self,
        document_id: str = "",
        tenant_id: str = "default",
        category: str = "general",
        chunk_size: int = 800,
        chunk_overlap: int = 80,
        parent_chunk_size: int = 2000,
    ) -> List[ParentChunk]:
        sheets_text = self._extract_sheets()
        full_text = "\n\n".join(sheets_text)
        from app.rag.chunkers.smart_chunker import (
            extract_qa_pairs,
            merge_qa_into_chunks,
            split_markdown_by_heading,
        )
        sections = split_markdown_by_heading(full_text, chunk_size, chunk_overlap)
        qa_pairs = extract_qa_pairs(full_text)
        if qa_pairs:
            sections = merge_qa_into_chunks(sections, qa_pairs)
        child_chunks = build_child_chunks(
            sections,
            source_filename=self.filename,
            document_id=document_id,
            tenant_id=tenant_id,
            category=category,
        )
        return build_parent_chunks(
            child_chunks,
            parent_chunk_size=parent_chunk_size,
            source_filename=self.filename,
            document_id=document_id,
            tenant_id=tenant_id,
        )

    def _extract_sheets(self) -> List[str]:
        """提取每个工作表为 Markdown 表格（每行 1 切片）。"""
        try:
            from openpyxl import load_workbook
            import io
            binary = self._read_file()
            wb = load_workbook(io.BytesIO(binary), data_only=True)
            output = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                md_table = self._sheet_to_markdown(ws)
                if md_table:
                    output.append(f"## Sheet: {sheet_name}\n\n{md_table}")
            return output
        except ImportError:
            logger.warning("openpyxl 未安装")
            return []

    def _sheet_to_markdown(self, ws) -> str:
        """工作表转 Markdown 表格（展开合并单元格）。"""
        if ws.max_row == 0 or ws.max_column == 0:
            return ""
        # 处理合并单元格：把合并值广播到所有覆盖位置
        merge_map = {}
        for merged in ws.merged_cells.ranges:
            min_row, min_col, max_row, max_col = merged.min_row, merged.min_col, merged.max_row, merged.max_col
            top_left = ws.cell(row=min_row, column=min_col).value
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    merge_map[(r, c)] = top_left
        # 渲染
        lines = []
        for row in range(1, ws.max_row + 1):
            cells = []
            for col in range(1, ws.max_column + 1):
                v = merge_map.get((row, col), ws.cell(row=row, column=col).value)
                cells.append("" if v is None else str(v).replace("|", "\|").replace("\n", " "))
            lines.append("| " + " | ".join(cells) + " |")
        if not lines:
            return ""
        header_sep = "| " + " | ".join(["---"] * ws.max_column) + " |"
        return "\n".join([lines[0], header_sep] + lines[1:])

    def _read_file(self) -> bytes:
        if self.file_url:
            return download_file(self.file_url, MAX_FILE_SIZE)
        with open(self.file_path, "rb") as f:
            return f.read()
